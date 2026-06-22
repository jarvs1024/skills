#!/usr/bin/env python3
"""smoke test: 用一份合成 MD 跑 xlsx pipeline。
跑不通则 exit 1, 跑通则列出关键断言结果。
可以在 CI 或 pre-commit 里跑, 不依赖外部硬件。

注意: 这是**测试脚本**, 不是产线代码。正常运行 skill 不需要执行它。
仅在改完 parse_md / 列宽算法 / generate_xlsx 后跑一次, 确认无回归。
"""
import sys
import os
import pathlib
import tempfile
import traceback

# Windows 终端
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# 让 import 找到 sibling scripts
SCRIPTS_DIR = pathlib.Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))

# 动态导入 (避免在 skill 安装路径无 __init__.py 时 import 失败)
import importlib.util

def _load(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod

gen_mod = _load("generate_xlsx", SCRIPTS_DIR / "generate_xlsx.py")

# 合成 MD, 覆盖关键场景:
# - 多个工作模块, 每个模块多条子内容 (<br>)
# - 单条子内容 (不编号)
# - 含 <br> 换行
# - 含 issue / 版本号
# - 含空行/空表格
SAMPLE_MD = """# 周报 (2026-06-W25)

> 生成于 2026-06-19 23:30 · 本周工作日 5 天

## 本周工作总结

| 序号 | 工作模块 | 工作内容与成果 | 完成进度 | 风险点 |
|---|---|---|---|---|
| 1 | SSD 老化与回归 | 1) 写 fio 随机写 latency 脚本,支持顺序/随机/混合三种负载<br>2) 跑 3 块样品 P99 latency 回归,发现 2 个异常值;提 issue #412、#413<br>3) 老化机 #7 升级 FW 完成后回归通过<br>4) 跟硬件组对 FW trim 状态机问题 | 进行中 | 等开发定位 #412 根因 |
| 2 | 研发协作 | review 小李 PR #218,提 4 条建议(变量命名、异常处理等) | 已完成 |  |
| 3 | 准入准出文档 | 推进《老化准入准出》v0.3 起草,补全异常场景清单 | 进行中 |  |

## 下周工作计划

| 序号 | 优先级 | 工作计划 |
|---|---|---|
| 1 | P0 | 推进 BIB_V1.4.3 154 转测适配,继续 bics8 兼容性测试 |
| 2 | P1 | 修复 #413 回归脚本报错,跑完 P99 latency 全量回归 |
"""

def assert_eq(actual, expected, label):
    if actual != expected:
        print(f"  [FAIL] {label}: 期望 {expected!r}, 实际 {actual!r}")
        return False
    print(f"  [OK]   {label}")
    return True

def assert_true(cond, label):
    if not cond:
        print(f"  [FAIL] {label}")
        return False
    print(f"  [OK]   {label}")
    return True

def main():
    print("=" * 60)
    print("weekly-report smoke test")
    print("=" * 60)
    passed = 0
    failed = 0
    # --- 1. parse_md ---
    print("\n[1] parse_md")
    try:
        with tempfile.TemporaryDirectory() as tmp:
            md_path = pathlib.Path(tmp) / "sample.md"
            md_path.write_text(SAMPLE_MD, encoding="utf-8")
            meta, tables = gen_mod.parse_md(md_path)
            if assert_true(len(tables) == 2, "解析出 2 个表"):
                passed += 1
            else:
                failed += 1
            if assert_true("生成于" in meta, "meta 包含 '生成于'"):
                passed += 1
            else:
                failed += 1
            # 新增: md 必须含两个 H2 小标题, 跟 xlsx sheet 标题对齐
            raw_md = md_path.read_text(encoding="utf-8")
            if assert_true("## 本周工作总结" in raw_md, "md 含 ## 本周工作总结 小标题"):
                passed += 1
            else:
                failed += 1
            if assert_true("## 下周工作计划" in raw_md, "md 含 ## 下周工作计划 小标题"):
                passed += 1
            else:
                failed += 1
            header, body = tables[0]
            if assert_eq(len(header), 5, "summary header 列数 = 5"):
                passed += 1
            else:
                failed += 1
            if assert_eq(len(body), 3, "summary body 行数 = 3"):
                passed += 1
            else:
                failed += 1
            # 关键断言: <br> 已转 \n
            cell_c = body[0][2]
            if assert_true("\n" in cell_c, "<br> 已被 parse_md 转成 \\n"):
                passed += 1
            else:
                failed += 1
            if assert_true("<br>" not in cell_c, "cell 中无残留 <br>"):
                passed += 1
            else:
                failed += 1
            if assert_eq(cell_c.count("\n"), 3, "4 条子条目 = 3 个 \\n"):
                passed += 1
            else:
                failed += 1
            # --- 2. md_to_xlsx ---
            print("\n[2] md_to_xlsx")
            xlsx_path = pathlib.Path(tmp) / "sample.xlsx"
            gen_mod.md_to_xlsx(md_path, xlsx_path, today=__import__("datetime").date(2026, 6, 18))
            if assert_true(xlsx_path.exists(), "xlsx 文件已生成"):
                passed += 1
            else:
                failed += 1
            # 验证 xlsx 内容
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb.active
            # 行 4 是第一条数据 (meta=1, title=2, header=3, data=4)
            cell = ws.cell(4, 3).value or ""
            if assert_true("\n" in cell, "xlsx cell 含真换行"):
                passed += 1
            else:
                failed += 1
            if assert_true("<br>" not in cell, "xlsx cell 无 <br> 字面量"):
                passed += 1
            else:
                failed += 1
            # 列宽
            col_c_width = ws.column_dimensions["C"].width
            if assert_true(col_c_width and col_c_width > 50, f"工作内容列宽 {col_c_width} > 50"):
                passed += 1
            else:
                failed += 1
                failed += 1
    except Exception:
        print("[CRASH]")
        traceback.print_exc()
        failed += 1
    print()
    print("=" * 60)
    print(f"通过 {passed} / 失败 {failed}")
    print("=" * 60)
    sys.exit(0 if failed == 0 else 1)

if __name__ == "__main__":
    main()
