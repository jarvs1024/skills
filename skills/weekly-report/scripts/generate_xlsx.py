#!/usr/bin/env python3
"""Markdown 周报 → 公司风 Excel。
特性:
- 列宽按内容**最长单行**算 (取 \\n 切分后的最大, 不取整段)
- 内容列自动换行 + 左对齐
- 下周表头"工作计划"格子横跨剩余列
- 标题行也横跨 A:E
- 本周范围 = 上周五 ~ 本周四 (5 个工作日跨周末)
- 下周范围 = 本周五 ~ 下周四
- 兼容 Windows / macOS / Linux (UTF-8 stdout, pathlib 路径)
"""
import sys
import io
import os
import pathlib
from datetime import date, timedelta

# 兼容早期 Python (openpyxl 通常装在 3.8+)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "[ERROR] 缺 openpyxl: pip3 install --user openpyxl\n"
    )
    sys.exit(2)

# Windows 终端默认 GBK, print 中文会崩
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# === 配色 ===
META_FILL    = PatternFill("solid", fgColor="F2F2F2")
TITLE_FILL   = PatternFill("solid", fgColor="8FAADC")
HEADER_FILL  = PatternFill("solid", fgColor="D9E2F3")
DATA_FILL    = PatternFill("solid", fgColor="FFFFFF")
EMPTY_FILL   = PatternFill("solid", fgColor="FFFFFF")

# 字体: 优先 Microsoft YaHei, fallback 微软雅黑
META_FONT    = Font(name="Microsoft YaHei", size=10, color="333333")
TITLE_FONT   = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
HEADER_FONT  = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
NORMAL_FONT  = Font(name="Microsoft YaHei", size=11, color="1A1A1A")

# 对齐
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP    = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

# 边框
THIN_GREY    = Side(border_style="thin", color="D0D0D0")
GREY_BORDER  = Border(top=THIN_GREY, bottom=THIN_GREY, left=THIN_GREY, right=THIN_GREY)

# 行高
ROW_H = {"meta": 22, "title": 36, "header": 28, "data": 36}

def report_window(today: date):
    """周四发送口径下的"本周"和"下周"日期范围。
    本周 = 上周五 ~ 本周四
    下周 = 本周五 ~ 下周四
    """
    if today.weekday() == 3:  # 周四
        this_thursday = today
    else:
        days_back = (today.weekday() - 3) % 7
        this_thursday = today - timedelta(days=days_back)
    last_friday = this_thursday - timedelta(days=6)
    this_friday = this_thursday + timedelta(days=1)
    next_thursday = this_thursday + timedelta(days=7)
    return (last_friday, this_thursday), (this_friday, next_thursday)

def fmt_date(d: date):
    return f"{d.year}.{d.month}.{d.day}"

def fmt_range(start: date, end: date):
    if start.year == end.year and start.month == end.month:
        return f"{fmt_date(start)} ~ {end.day}"
    if start.year == end.year:
        return f"{start.year}.{start.month}.{start.day} ~ {end.month}.{end.day}"
    return f"{fmt_date(start)} ~ {fmt_date(end)}"

def build_titles(today: date):
    (this_s, this_e), (next_s, next_e) = report_window(today)
    return {
        "summary": f"本周工作总结({fmt_range(this_s, this_e)})",
        "plan":    f"下周工作计划({fmt_range(next_s, next_e)})",
    }

def parse_md(md_path):
    """读 markdown, 提取 meta 行 + 表格列表。
    表格要求: 至少 2 行 (header + 1 body)
    返回: (meta_text, [(header_row, [body_row, body_row, ...]), ...])
    """
    md_path = pathlib.Path(md_path)
    if not md_path.exists():
        raise FileNotFoundError(f"MD 文件不存在: {md_path}")
    text = md_path.read_text(encoding="utf-8")
    # 收集所有 meta 行 (> 开头)
    meta_lines = []
    for ln in text.splitlines():
        if ln.strip().startswith("> "):
            meta_lines.append(ln.strip()[2:].strip())
        elif ln.strip() and not ln.strip().startswith(">"):
            if meta_lines:  # 已收过 meta, 遇到非空非 > 行就停
                break
    meta = " | ".join(meta_lines)
    # 解析表格
    tables = []
    current, in_table = [], False
    for ln in text.splitlines():
        s = ln.strip()
        if s.startswith("|"):
            current.append([c.strip() for c in s.strip("|").split("|")])
            in_table = True
        else:
            if current and in_table:
                tables.append(current)
                current, in_table = [], False
    if current:
        tables.append(current)
    # 过滤掉分隔行 (| --- | --- |)
    def is_separator(row):
        if not row:
            return False
        return all(set(c.replace(" ", "")) <= set("-:") for c in row if c)
    cleaned = []
    for t in tables:
        if len(t) < 2:
            continue
        # 跳过全部分隔行
        body = [r for r in t[1:] if not is_separator(r)]
        if not body:
            continue
        header = t[0]
        # 转 <br> → 真换行
        def _br(row):
            return [c.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n") for c in row]
        cleaned.append((_br(header), [_br(r) for r in body]))
    return meta, cleaned

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        if w is not None:
            ws.column_dimensions[get_column_letter(i)].width = w

def _paint_row(ws, row, n_cols, fill, font, alignment, height):
    for j in range(1, n_cols + 1):
        c = ws.cell(row=row, column=j)
        c.fill = fill
        c.font = font
        c.alignment = alignment
        c.border = GREY_BORDER
    ws.row_dimensions[row].height = height

def write_meta_row(ws, meta_text, n_cols, row, widths):
    ws.cell(row=row, column=1, value=meta_text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    _paint_row(ws, row, n_cols, META_FILL, META_FONT, LEFT_WRAP, ROW_H["meta"])
    set_col_widths(ws, widths)
    return row + 1

def write_title_row(ws, title, n_cols, row, widths):
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    _paint_row(ws, row, n_cols, TITLE_FILL, TITLE_FONT, CENTER, ROW_H["title"])
    set_col_widths(ws, widths)
    return row + 1

def _display_len(s):
    """字符宽度: ASCII=1, 其它(CJK/全角)=2. 接近 Excel 列宽单位。"""
    length = 0
    for ch in str(s):
        if ord(ch) > 127:
            length += 2
        else:
            length += 1
    return length

def _max_line_len(cell_text):
    """取 cell 内容按 \\n 切分后的最长单行长度。
    这样多子条目 (<br> 拆行) 时, 列宽按最长那一行算, 而不是整段字符数。
    """
    if cell_text is None:
        return 0
    s = str(cell_text)
    lines = s.split("\n")
    return max((_display_len(ln) for ln in lines), default=0)

def calc_widths(header, rows, default_widths):
    """按列计算宽度, 取该列所有 cell 的最长单行 + 2 字符 padding, 上限 100。"""
    widths = list(default_widths)
    for col_idx, default in enumerate(widths):
        max_len = _display_len(str(header[col_idx])) if col_idx < len(header) else 0
        for r in rows:
            if col_idx < len(r):
                max_len = max(max_len, _max_line_len(r[col_idx]))
        widths[col_idx] = max(default, min(max_len + 2, 100))
    return widths

def write_summary_table(ws, header, rows, widths, start_row):
    actual_widths = calc_widths(header, rows, widths)
    n_cols = len(header)
    for j, h in enumerate(header, 1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = GREY_BORDER
    ws.row_dimensions[start_row].height = ROW_H["header"]
    for i, row in enumerate(rows, 1):
        r = start_row + i
        for j in range(1, n_cols + 1):
            c = ws.cell(row=r, column=j)
            v = row[j-1] if j-1 < len(row) else None
            c.value = v
            c.fill = DATA_FILL
            c.font = NORMAL_FONT
            c.alignment = CENTER if j == 1 else LEFT_WRAP
            c.border = GREY_BORDER
        # 多行内容自动加高
        max_lines = 1
        for v in row:
            if v and "\n" in str(v):
                max_lines = max(max_lines, str(v).count("\n") + 1)
        ws.row_dimensions[r].height = ROW_H["data"] * max_lines
    set_col_widths(ws, actual_widths)
    return start_row + len(rows), actual_widths

def write_plan_table(ws, header, rows, widths, start_row):
    seq_header = header[0] if len(header) > 0 else "序号"
    prio_header = header[1] if len(header) > 1 else "优先级"
    plan_header = header[2] if len(header) > 2 else "工作计划"
    # 防御: rows 为空时只画表头
    c = ws.cell(row=start_row, column=1, value=seq_header)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = GREY_BORDER
    c = ws.cell(row=start_row, column=2, value=prio_header)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = GREY_BORDER
    ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=5)
    c = ws.cell(row=start_row, column=3, value=plan_header)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = GREY_BORDER
    for j in [4, 5]:
        cc = ws.cell(row=start_row, column=j)
        cc.fill = HEADER_FILL; cc.border = GREY_BORDER
    ws.row_dimensions[start_row].height = ROW_H["header"]
    for i, row in enumerate(rows, 1):
        r = start_row + i
        seq_val = row[0] if len(row) > 0 else None
        prio_val = row[1] if len(row) > 1 else None
        plan_val = row[2] if len(row) > 2 else None
        c = ws.cell(row=r, column=1, value=seq_val)
        c.fill = DATA_FILL; c.font = NORMAL_FONT; c.alignment = CENTER; c.border = GREY_BORDER
        c = ws.cell(row=r, column=2, value=prio_val)
        c.fill = DATA_FILL; c.font = NORMAL_FONT; c.alignment = CENTER; c.border = GREY_BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        c = ws.cell(row=r, column=3, value=plan_val)
        c.fill = DATA_FILL; c.font = NORMAL_FONT; c.alignment = LEFT_WRAP; c.border = GREY_BORDER
        for j in [4, 5]:
            cc = ws.cell(row=r, column=j)
            cc.fill = DATA_FILL; cc.border = GREY_BORDER
        # 多行内容自动加高
        max_lines = 1
        for v in row:
            if v and "\n" in str(v):
                max_lines = max(max_lines, str(v).count("\n") + 1)
        ws.row_dimensions[r].height = ROW_H["data"] * max_lines
    # 只调自己关心的列: 优先级(2)、工作计划(3), 避免污染 summary 列宽
    prio_col_max = max(_display_len("优先级"), _display_len("P0"), _display_len("P1"), _display_len("P2"))
    ws.column_dimensions[get_column_letter(2)].width = max(widths[1] if len(widths) > 1 else 10, prio_col_max + 2)
    plan_col_max = _max_line_len(plan_header)
    for row in rows:
        if len(row) > 2 and row[2]:
            plan_col_max = max(plan_col_max, _max_line_len(row[2]))
    ws.column_dimensions[get_column_letter(3)].width = min(plan_col_max + 2, 100)
    return start_row + max(len(rows), 1), widths

def md_to_xlsx(md_path, xlsx_path, today=None):
    md_path = pathlib.Path(md_path)
    xlsx_path = pathlib.Path(xlsx_path)
    if today is None:
        today = date.today()
    meta, tables = parse_md(md_path)
    if len(tables) < 2:
        raise ValueError(
            f"周报 MD 至少需要 2 个表 (本周总结 + 下周计划), 实际 {len(tables)} 个\n"
            f"  文件: {md_path}"
        )
    # 防御: 第二个表为空时插入 "(空)" 占位
    plan_table = tables[1] if len(tables[1]) > 0 else ([], [["", "", "(待填写)"]])
    titles = build_titles(today)
    # 确保输出目录存在
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "周报"
    SUMMARY_DEFAULT = [8, 18, 50, 14, 28]
    cur = 1
    cur = write_meta_row(ws, meta, n_cols=5, row=cur, widths=SUMMARY_DEFAULT)
    cur = write_title_row(ws, titles["summary"], n_cols=5, row=cur, widths=SUMMARY_DEFAULT)
    cur, w = write_summary_table(ws, *tables[0], widths=SUMMARY_DEFAULT, start_row=cur)
    cur += 1
    cur = write_title_row(ws, titles["plan"], n_cols=5, row=cur, widths=w)
    write_plan_table(ws, *plan_table, widths=w, start_row=cur)
    wb.save(xlsx_path)
    print(f"XLSX -> {xlsx_path}")
    print(f"  本周: {titles['summary']}")
    print(f"  下周: {titles['plan']}")

def _parse_date_arg(s):
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        raise SystemExit(
            f"[ERROR] --date 格式错误: {s!r} (期望 YYYY-MM-DD, 例如 2026-06-18)"
        )

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Markdown 周报 → 公司风 Excel (.xlsx)",
        epilog="示例: python3 generate_xlsx.py 周报.md 周报.xlsx --date 2026-06-18",
    )
    parser.add_argument("md_path", help="输入的 markdown 路径")
    parser.add_argument("xlsx_path", help="输出的 xlsx 路径")
    parser.add_argument(
        "--date", type=str, default=None,
        help="本周四日期 YYYY-MM-DD (不传则用今天, 适用于补发)",
    )
    args = parser.parse_args()
    md_to_xlsx(args.md_path, args.xlsx_path, today=_parse_date_arg(args.date))
